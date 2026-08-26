#!/usr/bin/env python3
"""
Add "Robots" and "Follow Status" columns to a sitemap xlsx.

Reads the URL column of every sheet, fetches each page, and records the
meta robots / X-Robots-Tag directive.

    pip install requests openpyxl
    python3 add_robots_column.py sitemap_links.xlsx

Results are cached to <output>.cache.json after every batch, so you can Ctrl-C
and re-run without losing work — already-fetched URLs are skipped instantly.
"""

import argparse
import json
import os
import re
import sys
import time
import concurrent.futures
from copy import copy

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PAGE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}
MAX_HTML_BYTES = 250_000
SKIP_TYPES = ("image/", "video/", "audio/", "font/", "application/pdf",
              "application/zip")
CACHE_EVERY = 200

META_TAG_RE = re.compile(r"<meta\b[^>]*>", re.IGNORECASE)
ATTR_RE = re.compile(r"""([\w:-]+)\s*=\s*(?:"([^"]*)"|'([^']*)'|([^\s"'>]+))""")


def parse_meta_robots(html):
    """Content of <meta name="robots">, falling back to name="googlebot"."""
    found = {}
    for tag in META_TAG_RE.finditer(html):
        attrs = {}
        for m in ATTR_RE.finditer(tag.group(0)):
            attrs[m.group(1).lower()] = m.group(2) or m.group(3) or m.group(4) or ""
        name = attrs.get("name", "").strip().lower()
        if name in ("robots", "googlebot") and name not in found:
            found[name] = attrs.get("content", "").strip()
    return found.get("robots") or found.get("googlebot") or ""


def fetch_directive(session, url, timeout):
    try:
        resp = session.get(url, headers=PAGE_HEADERS, timeout=timeout,
                           stream=True, allow_redirects=True)
    except requests.RequestException as e:
        return f"Error: {type(e).__name__}"

    try:
        if resp.status_code != 200:
            return f"HTTP {resp.status_code}"

        header_val = resp.headers.get("X-Robots-Tag", "").strip()

        meta_val = ""
        ctype = resp.headers.get("Content-Type", "").lower()
        if not any(t in ctype for t in SKIP_TYPES):
            size, chunks = 0, []
            for chunk in resp.iter_content(8192):
                chunks.append(chunk)
                size += len(chunk)
                if size >= MAX_HTML_BYTES:
                    break
            html = b"".join(chunks).decode("utf-8", errors="ignore")
            head = re.split(r"</head\s*>", html, maxsplit=1, flags=re.IGNORECASE)[0]
            meta_val = parse_meta_robots(head)
    finally:
        resp.close()

    if meta_val and header_val:
        return f"{meta_val} (X-Robots-Tag: {header_val})"
    if meta_val:
        return meta_val
    if header_val:
        return f"X-Robots-Tag: {header_val}"
    return "Not set"


def derive_status(directive):
    """(Index Status, Follow Status) from a raw directive string."""
    d = (directive or "").lower()
    if d.startswith(("http ", "error", "not checked")):
        return "Unknown", "Unknown"
    if d == "not set" or not d:
        return "Index (default)", "Follow (default)"
    is_none = bool(re.search(r"\bnone\b", d))
    index = "Noindex" if ("noindex" in d or is_none) else "Index"
    follow = "Nofollow" if ("nofollow" in d or is_none) else "Follow"
    return index, follow


def make_session(workers):
    s = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=workers, pool_maxsize=workers, max_retries=0
    )
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def collect_urls(wb):
    """Every distinct URL across all sheets, plus each sheet's URL column index."""
    urls, url_cols = [], {}
    seen = set()
    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        try:
            col = headers.index("URL") + 1
        except ValueError:
            print(f"  ! {ws.title}: no 'URL' header, skipping")
            continue
        url_cols[ws.title] = col
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
            u = row[0]
            if u and u not in seen:
                seen.add(u)
                urls.append(u)
    return urls, url_cols


def crawl(urls, cache, cache_path, workers, timeout):
    todo = [u for u in urls if u not in cache]
    if not todo:
        print(f"All {len(urls)} URLs already cached.")
        return
    print(f"{len(urls)} URLs total, {len(cache)} cached, {len(todo)} to fetch "
          f"({workers} workers)")

    session = make_session(workers)
    done, start = 0, time.time()
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
            futures = {ex.submit(fetch_directive, session, u, timeout): u for u in todo}
            for fut in concurrent.futures.as_completed(futures):
                url = futures[fut]
                try:
                    cache[url] = fut.result()
                except Exception as e:
                    cache[url] = f"Error: {type(e).__name__}"
                done += 1
                if done % CACHE_EVERY == 0 or done == len(todo):
                    save_cache(cache, cache_path)
                    rate = done / max(time.time() - start, 0.001)
                    eta = (len(todo) - done) / max(rate, 0.001)
                    print(f"  {done}/{len(todo)}  {rate:.1f}/s  eta {eta/60:.1f}m",
                          flush=True)
    except KeyboardInterrupt:
        save_cache(cache, cache_path)
        print(f"\nInterrupted. {len(cache)} results cached — re-run to resume.")
        sys.exit(1)
    finally:
        session.close()
        save_cache(cache, cache_path)


def save_cache(cache, path):
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(cache, f)
    os.replace(tmp, path)


def write_columns(wb, url_cols, cache, add_index):
    """Insert Robots (+ Index Status) + Follow Status right after the URL column."""
    new_cols = ["Robots", "Index Status", "Follow Status"] if add_index \
        else ["Robots", "Follow Status"]

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    CELL_FONT = Font(name="Arial", size=10)
    BAD_FONT = Font(name="Arial", size=10, color="C00000", bold=True)
    WARN_FONT = Font(name="Arial", size=10, color="808080", italic=True)

    counts = {}
    for ws in wb.worksheets:
        col = url_cols.get(ws.title)
        if not col:
            continue

        ws.insert_cols(col + 1, amount=len(new_cols))

        hdr_style = ws.cell(row=1, column=col)
        for i, name in enumerate(new_cols):
            c = ws.cell(row=1, column=col + 1 + i, value=name)
            c.font = HEADER_FONT
            c.fill = copy(hdr_style.fill)
            c.alignment = copy(hdr_style.alignment)

        for r in range(2, ws.max_row + 1):
            url = ws.cell(row=r, column=col).value
            if not url:
                continue
            directive = cache.get(url, "Not checked")
            idx, fol = derive_status(directive)
            values = [directive, idx, fol] if add_index else [directive, fol]
            counts[fol] = counts.get(fol, 0) + 1
            for i, v in enumerate(values):
                c = ws.cell(row=r, column=col + 1 + i, value=v)
                low = str(v).lower()
                if "nofollow" in low or "noindex" in low or re.search(r"\bnone\b", low):
                    c.font = BAD_FONT
                elif low.startswith(("error", "http ", "unknown", "not checked")):
                    c.font = WARN_FONT
                else:
                    c.font = CELL_FONT

        ws.column_dimensions[get_column_letter(col + 1)].width = 30
        for i in range(1, len(new_cols)):
            ws.column_dimensions[get_column_letter(col + 1 + i)].width = 16

        last = get_column_letter(ws.max_column)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{last}{ws.max_row}"

    return counts


def main():
    p = argparse.ArgumentParser()
    p.add_argument("input")
    p.add_argument("-o", "--output")
    p.add_argument("-w", "--workers", type=int, default=16)
    p.add_argument("-t", "--timeout", type=int, default=15)
    p.add_argument("--cache")
    p.add_argument("--index-column", action="store_true",
                   help="also add an Index/Noindex column")
    args = p.parse_args()

    out = args.output or args.input.replace(".xlsx", "_robots.xlsx")
    cache_path = args.cache or out + ".cache.json"

    cache = {}
    if os.path.exists(cache_path):
        with open(cache_path) as f:
            cache = json.load(f)

    print(f"Reading {args.input}")
    wb = load_workbook(args.input)
    urls, url_cols = collect_urls(wb)

    crawl(urls, cache, cache_path, args.workers, args.timeout)

    counts = write_columns(wb, url_cols, cache, args.index_column)
    wb.save(out)

    print(f"\nSaved {out}")
    for k in sorted(counts):
        print(f"  {k:<18} {counts[k]}")


if __name__ == "__main__":
    main()
