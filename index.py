from fastapi import FastAPI, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
import requests
import re
from lxml import etree
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

app = FastAPI()

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; SitemapScraper/1.0)"}
NS = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
LANG_NAMES = {
    "en": "English", "fr": "French", "de": "German", "es": "Spanish",
    "zh": "Chinese", "ko": "Korean", "pt": "Portuguese", "ru": "Russian",
    "tr": "Turkish", "ja": "Japanese", "it": "Italian", "nl": "Dutch"
}

class AnalyzeRequest(BaseModel):
    url: str

class ScrapeRequest(BaseModel):
    url: str
    categories: list[str]

def fetch_sitemap(url):
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    content = resp.content
    if url.endswith(".gz"):
        import gzip
        content = gzip.decompress(content)
    return content

def guess_type(url):
    filename = url.rstrip("/").split("/")[-1]
    match = re.match(r"sitemap[_-]?(.+)\.xml", filename, re.IGNORECASE)
    if match:
        raw = match.group(1).strip("-_ ")
        return raw.replace("-", " ").replace("_", " ").title()
    return filename

def guess_lang(url):
    path = urlparse(url).path
    lang_match = re.match(r"^/([a-z]{2})/", path)
    return lang_match.group(1) if lang_match else "en"

def extract_urls(url):
    content = fetch_sitemap(url)
    tree = etree.fromstring(content)
    entries = []
    for u in tree.findall("sm:url", NS):
        loc = u.find("sm:loc", NS)
        lastmod = u.find("sm:lastmod", NS)
        changefreq = u.find("sm:changefreq", NS)
        priority = u.find("sm:priority", NS)
        entries.append({
            "URL": loc.text.strip() if loc is not None else "",
            "Last Modified": lastmod.text.strip() if lastmod is not None else "",
            "Change Freq": changefreq.text.strip() if changefreq is not None else "",
            "Priority": priority.text.strip() if priority is not None else "",
        })
    return entries

def add_sheet(wb, name, rows):
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    CELL_FONT = Font(name="Arial", size=10)
    THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
    HDR = ["URL", "Last Modified", "Change Freq", "Priority"]

    name = name[:31]
    ws = wb.create_sheet(title=name)
    for ci, h in enumerate(HDR, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, HEADER_ALIGN
    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(HDR, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font, c.border = CELL_FONT, THIN_BORDER
    for col, w in {"A": 80, "B": 22, "C": 14, "D": 10}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

def get_categories(sitemap_url):
    content = fetch_sitemap(sitemap_url)
    index_tree = etree.fromstring(content)
    root_tag = etree.QName(index_tree.tag).localname

    if root_tag == "sitemapindex":
        child_locs = [loc.text.strip() for loc in index_tree.findall("sm:sitemap/sm:loc", NS)]
    else:
        child_locs = [sitemap_url]

    categories = {}
    for child_url in child_locs:
        stype = guess_type(child_url)
        lang = guess_lang(child_url)
        categories.setdefault(stype, []).append({"url": child_url, "lang": lang})
    return categories

@app.post("/api/analyze")
def analyze(req: AnalyzeRequest):
    try:
        categories = get_categories(req.url)
        
        # Format for frontend
        response_data = []
        for stype, items in categories.items():
            langs = sorted(set(item["lang"] for item in items))
            lang_labels = ", ".join(LANG_NAMES.get(l, l.upper()) for l in langs)
            response_data.append({
                "type": stype,
                "count": len(items),
                "langs": lang_labels
            })
        return {"categories": response_data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/scrape")
def scrape(req: ScrapeRequest):
    try:
        all_categories = get_categories(req.url)
        scraped_data = {}
        
        for stype in req.categories:
            if stype in all_categories:
                for item in all_categories[stype]:
                    urls = extract_urls(item["url"])
                    scraped_data[(stype, item["lang"])] = urls

        wb = Workbook()
        wb.remove(wb.active)

        types_in_data = sorted(set(t for t, l in scraped_data.keys()))
        for stype in types_in_data:
            langs = sorted([l for (t, l) in scraped_data if t == stype])
            for lang in langs:
                lang_label = LANG_NAMES.get(lang, lang.upper())
                sheet_name = stype if len(langs) == 1 and lang == "en" else f"{stype} - {lang_label}"
                add_sheet(wb, sheet_name, scraped_data[(stype, lang)])

        output = BytesIO()
        wb.save(output)
        excel_data = output.getvalue()

        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=sitemap_links.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))