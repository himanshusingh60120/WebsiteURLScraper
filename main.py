from fastapi import FastAPI, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
import requests
import re
from lxml import etree
from lxml import html as lxml_html
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor

app = FastAPI()

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; SitemapScraper/1.0)"}
NS = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

LANG_NAMES = {
    "en": "English", "fr": "French", "de": "German", "es": "Spanish",
    "zh": "Chinese", "ko": "Korean", "pt": "Portuguese", "ru": "Russian",
    "tr": "Turkish", "ja": "Japanese", "it": "Italian", "nl": "Dutch"
}

# Optional per-page fields: key sent by the frontend -> Excel column header
FIELD_LABELS = {
    "meta_title": "Meta Title",
    "meta_description": "Meta Description",
    "h1": "H1",
}
FIELD_ORDER = ["meta_title", "meta_description", "h1"]
SITEMAP_COLUMNS = ["Last Modified", "Change Freq", "Priority"]

COL_WIDTHS = {
    "URL": 70,
    "Meta Title": 55,
    "Meta Description": 75,
    "H1": 45,
    "Status": 10,
    "Last Modified": 22,
    "Change Freq": 14,
    "Priority": 10,
}

MAX_WORKERS = 20          # parallel page fetches inside one batch
PAGE_TIMEOUT = 10         # seconds per page
MAX_HTML_BYTES = 300_000  # stop reading a page after this much HTML
MAX_BATCH = 120           # hard cap on URLs accepted per /api/pages call

SESSION = requests.Session()
SESSION.mount("https://", requests.adapters.HTTPAdapter(pool_maxsize=MAX_WORKERS))
SESSION.mount("http://", requests.adapters.HTTPAdapter(pool_maxsize=MAX_WORKERS))


class AnalyzeRequest(BaseModel):
    url: str


class UrlsRequest(BaseModel):
    url: str
    categories: list[str]


class PagesRequest(BaseModel):
    urls: list[str]
    fields: list[str] = []


class ExportSheet(BaseModel):
    name: str
    rows: list[dict]


class ExportRequest(BaseModel):
    sheets: list[ExportSheet]
    fields: list[str] = []


# ---------------------------------------------------------------------------
# Sitemap parsing
# ---------------------------------------------------------------------------

def fetch_sitemap(url):
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    content = resp.content
    if url.endswith(".gz"):
        import gzip
        content = gzip.decompress(content)
    return content


def guess_type(url):
    filename = url.rstrip("/").split("/")[-1]
    match = re.match(r"sitemap[_-]?(.+)\.xml", filename, re.IGNORECASE)
    if match:
        raw = match.group(1).strip("-_ ")
        return raw.replace("-", " ").replace("_", " ").title()
    return filename


def guess_lang(url):
    path = urlparse(url).path
    lang_match = re.match(r"^/([a-z]{2})/", path)
    return lang_match.group(1) if lang_match else "en"


def extract_urls(url):
    content = fetch_sitemap(url)
    tree = etree.fromstring(content)
    entries = []
    for u in tree.findall("sm:url", NS):
        loc = u.find("sm:loc", NS)
        lastmod = u.find("sm:lastmod", NS)
        changefreq = u.find("sm:changefreq", NS)
        priority = u.find("sm:priority", NS)
        entries.append({
            "URL": loc.text.strip() if loc is not None else "",
            "Last Modified": lastmod.text.strip() if lastmod is not None else "",
            "Change Freq": changefreq.text.strip() if changefreq is not None else "",
            "Priority": priority.text.strip() if priority is not None else "",
        })
    return entries


def get_categories(sitemap_url):
    content = fetch_sitemap(sitemap_url)
    index_tree = etree.fromstring(content)
    root_tag = etree.QName(index_tree.tag).localname

    if root_tag == "sitemapindex":
        child_locs = [loc.text.strip() for loc in index_tree.findall("sm:sitemap/sm:loc", NS)]
    else:
        child_locs = [sitemap_url]

    categories = {}
    for child_url in child_locs:
        stype = guess_type(child_url)
        lang = guess_lang(child_url)
        categories.setdefault(stype, []).append({"url": child_url, "lang": lang})
    return categories


# ---------------------------------------------------------------------------
# Page-level extraction (meta title / meta description / H1)
# ---------------------------------------------------------------------------

def clean(text):
    return re.sub(r"\s+", " ", (text or "")).strip()


def first_value(values):
    for value in values:
        cleaned = clean(value)
        if cleaned:
            return cleaned
    return ""


def parse_page_fields(html_bytes, fields):
    data = {FIELD_LABELS[f]: "" for f in fields}
    if not html_bytes:
        return data

    doc = lxml_html.fromstring(html_bytes)

    if "meta_title" in fields:
        title = first_value(doc.xpath("//head/title//text()"))
        if not title:
            title = first_value(doc.xpath("//title//text()"))
        if not title:
            title = first_value(doc.xpath("//meta[@property='og:title']/@content"))
        data["Meta Title"] = title

    if "meta_description" in fields:
        desc = first_value(doc.xpath(
            "//meta[translate(@name, 'DESCRIPTION', 'description')='description']/@content"
        ))
        if not desc:
            desc = first_value(doc.xpath("//meta[@property='og:description']/@content"))
        data["Meta Description"] = desc

    if "h1" in fields:
        headings = [clean(h.text_content()) for h in doc.xpath("//h1")]
        data["H1"] = " | ".join(h for h in headings if h)

    return data


def fetch_page(url, fields):
    result = {FIELD_LABELS[f]: "" for f in fields}
    result["Status"] = ""

    if not url:
        result["Status"] = "No URL"
        return result

    try:
        resp = SESSION.get(
            url, headers=HEADERS, timeout=PAGE_TIMEOUT,
            stream=True, allow_redirects=True
        )
        result["Status"] = str(resp.status_code)

        content_type = resp.headers.get("Content-Type", "").lower()
        if resp.status_code != 200 or "html" not in content_type:
            resp.close()
            return result

        buffer = bytearray()
        for chunk in resp.iter_content(16384):
            buffer.extend(chunk)
            if len(buffer) >= MAX_HTML_BYTES:
                break
        resp.close()

        result.update(parse_page_fields(bytes(buffer), fields))
    except Exception as exc:
        result["Status"] = f"Error: {type(exc).__name__}"

    return result


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def build_headers(fields):
    headers = ["URL"]
    headers += [FIELD_LABELS[f] for f in FIELD_ORDER if f in fields]
    if fields:
        headers.append("Status")
    headers += SITEMAP_COLUMNS
    return headers


def add_sheet(wb, name, rows, headers):
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    CELL_FONT = Font(name="Arial", size=10)
    CELL_ALIGN = Alignment(vertical="top")
    THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

    name = re.sub(r"[\\/*?:\[\]]", "-", name)[:31] or "Sheet"
    ws = wb.create_sheet(title=name)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, HEADER_ALIGN

    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font, c.border, c.alignment = CELL_FONT, THIN_BORDER, CELL_ALIGN

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(h, 20)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.post("/api/analyze")
def analyze(req: AnalyzeRequest):
    """Step 1: list sitemap categories so the user can pick."""
    try:
        categories = get_categories(req.url)
        response_data = []
        for stype, items in categories.items():
            langs = sorted(set(item["lang"] for item in items))
            lang_labels = ", ".join(LANG_NAMES.get(l, l.upper()) for l in langs)
            response_data.append({
                "type": stype,
                "count": len(items),
                "langs": lang_labels
            })
        return {"categories": response_data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/urls")
def urls(req: UrlsRequest):
    """Step 2: return every URL, grouped into sheets. No page fetching, so it's fast."""
    try:
        all_categories = get_categories(req.url)
        collected = {}

        for stype in req.categories:
            for item in all_categories.get(stype, []):
                key = (stype, item["lang"])
                collected.setdefault(key, []).extend(extract_urls(item["url"]))

        sheets = []
        for stype in sorted(set(t for t, _ in collected)):
            langs = sorted(l for (t, l) in collected if t == stype)
            for lang in langs:
                lang_label = LANG_NAMES.get(lang, lang.upper())
                name = stype if len(langs) == 1 and lang == "en" else f"{stype} - {lang_label}"
                sheets.append({"name": name, "rows": collected[(stype, lang)]})

        total = sum(len(s["rows"]) for s in sheets)
        return {"sheets": sheets, "total": total}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/pages")
def pages(req: PagesRequest):
    """Step 3: fetch one batch of pages. The browser calls this repeatedly."""
    fields = [f for f in FIELD_ORDER if f in req.fields]
    if not fields:
        raise HTTPException(status_code=400, detail="No page fields requested")
    if len(req.urls) > MAX_BATCH:
        raise HTTPException(status_code=400, detail=f"Batch too large (max {MAX_BATCH})")

    if not req.urls:
        return {"results": []}

    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(req.urls))) as pool:
        results = list(pool.map(lambda u: fetch_page(u, fields), req.urls))

    return {"results": results}


@app.post("/api/export")
def export(req: ExportRequest):
    """Step 4: build the workbook from the rows the browser assembled."""
    try:
        fields = [f for f in FIELD_ORDER if f in req.fields]
        headers = build_headers(fields)

        wb = Workbook()
        wb.remove(wb.active)

        used = set()
        for sheet in req.sheets:
            name = sheet.name
            suffix = 2
            while name.lower() in used:
                name = f"{sheet.name[:27]} ({suffix})"
                suffix += 1
            used.add(name.lower())
            add_sheet(wb, name, sheet.rows, headers)

        if not wb.sheetnames:
            raise ValueError("Nothing to export")

        output = BytesIO()
        wb.save(output)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=sitemap_links.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
